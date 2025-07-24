import pandas as pd
import matplotlib.pyplot as plt
import openpyxl
from tabulate import tabulate

# Cargar el archivo Excel
archivo = './datos/top10s.xlsx'
df = pd.read_excel(archivo, sheet_name='top10s')
excel_dataframe = openpyxl.load_workbook('datos/top10s.xlsx')
dataframe = excel_dataframe.active

grafico = 'Si'  # Cambiar a 'Si' para generar gráfico
tabla = 'No'  # Cambiar a 'No' para no generar tabla

if grafico == 'Si':
    print("Generando gráfico.")
    print (df)

    # Análisis breve
    print("Forma del dataset:", df.shape)
    print("\nArtistas más comunes:")
    print(df['artist'].value_counts().head(10))

    print("\nPromedio de popularidad por año:")
    print(df.groupby('year')['pop'].mean())

    # Histograma de popularidad
    plt.figure(figsize=(8, 6))
    df['dur'].hist(bins=10, color='blue', edgecolor='black')
    plt.title('Histograma de popularidad de canciones')
    plt.xlabel('Popularidad')
    plt.ylabel('Frecuencia')
    plt.grid(True)
    plt.savefig('hist_popularidad.webp')
    plt.savefig('hist_popularidad.png')
    plt.show()

    if tabla == 'Si':
        print("Generando tabla...")
        # Generar tabla
        data = []

        for row in range(1, dataframe.max_row):
            # print(row)
            # _row = dataframe.cell(row, 1).value
            _row = [row, ]
            for col in dataframe.iter_cols(1, dataframe.max_column):
                # print(col[row].value)
                _row.append(col[row].value)
            data.append(_row)
        headers = ['#', 'title', 'artist', 'top genre', 'year', 'bpm', 'nrgy', 'dance', 'dB', 'live', 'val', 'dur',
                   'acous', 'spch', 'pop']
        headers_align = ['center'] * len(headers)
        print(tabulate(data, headers, tablefmt='fancy_grid', colalign='headers_align'))
    else:
        print("No se generará la tabla.")
else:
    print("No se generará el gráfico.")
    if tabla == 'Si':
        print("Generando tabla...")
        # Generar tabla
        data = []

        for row in range(1, dataframe.max_row):
            # print(row)
            #_row = dataframe.cell(row, 1).value
            _row = [row,]
            for col in dataframe.iter_cols(1, dataframe.max_column):
                # print(col[row].value)
                _row.append(col[row].value)
            data.append(_row)
        headers = ['#', 'title', 'artist', 'top genre', 'year', 'bpm', 'nrgy', 'dance', 'dB', 'live', 'val', 'dur', 'acous', 'spch', 'pop']
        headers_align = ['center'] * len(headers)
        print(tabulate(data,headers, tablefmt='fancy_grid', colalign='headers_align'))
    else:
        print("No se generará la tabla.")